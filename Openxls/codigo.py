import openpyxl
from openpyxl import Workbook
from datetime import datetime

def crear_archivo_excel(nombre_archivo):
    try:
        libro = openpyxl.load_workbook(nombre_archivo)
        print(f"El archivo {nombre_archivo} ya existe.")
    except FileNotFoundError:
        libro = Workbook()
        hoja = libro.active
        hoja.title = "Gastos"
        hoja.append(["Fecha", "Descripción", "Monto"])  # Encabezados
        libro.save(nombre_archivo)
        print(f"El archivo {nombre_archivo} ha sido creado.")

def agregar_gasto(nombre_archivo, fecha, descripcion, monto):
    libro = openpyxl.load_workbook(nombre_archivo)
    hoja = libro["Gastos"]
    hoja.append([fecha, descripcion, monto])
    libro.save(nombre_archivo)

def generar_informe(nombre_archivo):
    libro = openpyxl.load_workbook(nombre_archivo)
    hoja = libro["Gastos"]
    
    total_gastos = 0
    numero_gastos = 0
    gasto_mas_caro = ("", "", 0)  
    gasto_mas_barato = ("", "", float('inf'))
    
    for fila in hoja.iter_rows(min_row=2, values_only=True): 
        fecha, descripcion, monto = fila
        total_gastos += monto
        numero_gastos += 1
        
        if monto > gasto_mas_caro[2]:
            gasto_mas_caro = (fecha, descripcion, monto)
        
        if monto < gasto_mas_barato[2]:
            gasto_mas_barato = (fecha, descripcion, monto)
    
    print("\n--- Resumen de gastos ---")
    print(f"Número total de gastos: {numero_gastos}")
    print(f"Gasto más caro: {gasto_mas_caro[0]} - {gasto_mas_caro[1]} - ${gasto_mas_caro[2]:.2f}")
    print(f"Gasto más barato: {gasto_mas_barato[0]} - {gasto_mas_barato[1]} - ${gasto_mas_barato[2]:.2f}")
    print(f"Total gastado: ${total_gastos:.2f}")

def main():
    nombre_archivo = "informe_gastos.xlsx"
    crear_archivo_excel(nombre_archivo)
    
    while True:
        fecha_input = input("Introduce la fecha del gasto (YYYY-MM-DD): ")
        try:
            fecha = datetime.strptime(fecha_input, "%Y-%m-%d").date()
        except ValueError:
            print("Fecha no válida. Inténtalo de nuevo.")
            continue
        
        descripcion = input("Introduce la descripción del gasto: ")
        try:
            monto = float(input("Introduce el monto del gasto: "))
        except ValueError:
            print("Monto no válido. Inténtalo de nuevo.")
            continue
        
        agregar_gasto(nombre_archivo, fecha, descripcion, monto)
        
        continuar = input("¿Quieres añadir otro gasto? (s/n): ")
        if continuar.lower() != 's':
            break
    generar_informe(nombre_archivo)
    print(f"El informe ha sido guardado en {nombre_archivo}")

if __name__ == "__main__":
    main()
